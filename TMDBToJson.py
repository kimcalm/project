# 첫 데이터 베이스 세팅 (2004~2022년도 관객수 상위 30개씩에서 조금 더 필터링)

import requests
import json
from openpyxl import load_workbook

wb = load_workbook('./연도별 박스오피스 순위.xlsx')
ws = wb.active

# API_KEY='1c04f38f1d2f0979baea5787bc0cbdd8'
API_KEY='bd7f98121a9d0436318b3160e3374695'

def get_movie_datas():
  total_data = []

  for r in range(3, ws.max_row+1):
  # for r in range(3, 5):
    movie_title_en = ws.cell(row=r, column=6).value
    movie_title_kr = ws.cell(row=r, column=1).value
    nation = ws.cell(row=r, column=4).value
    
    url = f'https://api.themoviedb.org/3/search/movie?api_key={API_KEY}&language=ko-KR&page=1&include_adult=false&query={movie_title_en}'

    movies = requests.get(url).json()

    for movie in movies['results']:
      # print(year, type(year))
      if movie['poster_path'] and movie.get('release_date', '') and movie_title_kr == movie['title']:
        year = int(movie.get('release_date')[:4])
        if year >= 2004:
          print(movie_title_kr)
          fields = {
                      'title_en': movie_title_en,
                      'title_kr': movie['title'],
                      'nation': nation,
                      'released_date': movie['release_date'],
                      'genres': movie['genre_ids'],
                      'vote_avg': movie['vote_average'],
                      'overview': movie['overview'],
                      'poster_path': movie['poster_path'],
                    }
          
          data = {
                    "model": "movies.movie",
                    "pk": movie['id'],  
                    "fields": fields
                  }
          
          total_data.append(data)

  with open("movie_data.json", 'w', encoding='utf-8') as w:
    json.dump(total_data, w, indent=" ", ensure_ascii=False)

get_movie_datas()

