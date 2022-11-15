# kobis에서 찾은 데이터에 포스터 & 요약내용 추가하기

import requests
import json
from openpyxl import load_workbook

wb = load_workbook('./연도별 박스오피스 순위.xlsx')
ws = wb.active

API_KEY='bd7f98121a9d0436318b3160e3374695'

def get_movie_datas():
    total_data = []

    for r in range(4, ws.max_row+1):
    # for r in range(87, 88):
        movieCd = ws.cell(row=r, column=5).value
        movie_title_en = ws.cell(row=r, column=6).value
        year = ws.cell(row=r, column=2).value
        year = str(year)

        url = f'https://api.themoviedb.org/3/search/movie?api_key={API_KEY}&query={movie_title_en}&language=ko-KR'

        movies = requests.get(url).json()
        # print(movies)
        for movie in movies['results']:
            if movie.get('release_date'):
                tmdb_year = movie['release_date'][:4]
                    
                if tmdb_year == year:
                    print(movie_title_en)
                    fields = {
                            'adult': movie['adult'],
                            'vote_average': movie['vote_average'],
                            'overview': movie['overview'],
                            'poster_path': movie['poster_path'],
                            'movie_id': movieCd
                            }
                
                    data = {
                            "model": "movies.Tmdb_Movie",
                            "pk": movie['id'],  
                            "fields": fields
                            }
                
                    total_data.append(data)
                    break

    with open("tmdb_movie_data1.json", 'w', encoding='utf-8') as w:
        json.dump(total_data, w, indent=" ", ensure_ascii=False)

get_movie_datas()
