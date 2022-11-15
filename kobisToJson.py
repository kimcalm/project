# 

import requests
import json
from openpyxl import load_workbook

wb = load_workbook('./연도별 박스오피스 순위.xlsx')
ws = wb.active

API_KEY = 'a222b6703b63955f330a30797f941c12'

def get_movie_datas():
    total_data = []
    for r in range(331, ws.max_row+1):
    # for r in range(329, 330):
        movieCd = ws.cell(row=r, column=5).value
        url = f'http://www.kobis.or.kr/kobisopenapi/webservice/rest/movie/searchMovieInfo.json?key={API_KEY}&movieCd={movieCd}'

        movie = requests.get(url).json()
        movie = movie['movieInfoResult']['movieInfo']
        # print(movie)
        
        directors = []
        actors = []
        nations = []
        genres = []
        for direct in movie['directors']:
            directors.append(direct['peopleNm'])
        
        # print(directors)
        for act in movie['actors']:
            actors.append(act['peopleNm'])
        # print(actors)

        for genre in movie['genres']:
            genres.append(genre['genreNm'])

        for nation in movie['nations']:
            nations.append(nation['nationNm'])
        print(movie['movieNm'])
        fields = {
                    'movieNm': movie['movieNm'],
                    'movieNmEn': movie['movieNmEn'],
                    'showTm': movie['showTm'],
                    'prdtYear': movie['prdtYear'],
                    'nations': nations,
                    'genres': genres,
                    'directors': directors,
                    'actors': actors,
                }
        
        data = {
                "model": "movies.movie",
                "pk": int(movie['movieCd']),  
                "fields": fields
                }

        total_data.append(data)
    # print(total_data)

    with open("movie_data4.json", 'w', encoding='utf-8') as w:
        json.dump(total_data, w, indent=" ", ensure_ascii=False)

get_movie_datas()
